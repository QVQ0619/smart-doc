"""
基金项目申报规则抽取工具集

提供以下功能：
1. 多格式文档解析（PDF、Word、Excel、TXT、Markdown）
2. 文档结构提取（页码、章节信息）
3. 大模型 API 调用进行规则抽取
4. 结构化输出生成
"""

import os
import json
import re
from typing import Any, Dict, List, Optional, Annotated
from pathlib import Path

from langchain_core.tools import tool, InjectedToolArg

# 文档解析库
try:
    import fitz  # PyMuPDF for PDF
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from docx import Document
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    from openpyxl import load_workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


def _parse_pdf(file_path: str) -> List[Dict[str, Any]]:
    """
    解析 PDF 文件，返回带页码信息的段落列表
    
    Args:
        file_path: PDF 文件路径
        
    Returns:
        包含每页内容的列表，每项包含：
        - page_num: 页码（从 1 开始）
        - content: 文本内容
        - sections: 段落信息
    """
    if not HAS_PDF:
        raise ImportError("PyMuPDF (fitz) 未安装，无法解析 PDF 文件")
    
    result = []
    doc = fitz.open(file_path)
    
    for page_num, page in enumerate(doc, start=1):
        text = page.get_text()
        # 按段落分割
        paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
        
        result.append({
            "page_num": page_num,
            "content": text,
            "paragraphs": paragraphs,
            "filename": os.path.basename(file_path)
        })
    
    doc.close()
    return result


def _parse_docx(file_path: str) -> List[Dict[str, Any]]:
    """
    解析 Word 文档
    
    Args:
        file_path: Word 文件路径
        
    Returns:
        包含文档内容的列表（Word 无页码概念，用章节标识）
    """
    if not HAS_DOCX:
        raise ImportError("python-docx 未安装，无法解析 Word 文件")
    
    result = []
    doc = Document(file_path)
    
    # 提取段落和标题
    sections = []
    current_section = {"title": "正文", "content": []}
    
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
            
        # 检查是否是标题（通过样式或字体大小判断）
        if para.style.name.startswith('Heading') or len(text) < 50 and not any(c in text for c in '，。；：'):
            if current_section["content"]:
                sections.append(current_section)
            current_section = {"title": text, "content": []}
        else:
            current_section["content"].append(text)
    
    if current_section["content"]:
        sections.append(current_section)
    
    result.append({
        "section_num": 1,
        "content": "\n".join([s["title"] + "\n" + "\n".join(s["content"]) for s in sections]),
        "sections": sections,
        "filename": os.path.basename(file_path)
    })
    
    return result


def _parse_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """
    解析 Excel 文件
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        包含所有工作表内容的列表
    """
    if not HAS_XLSX:
        raise ImportError("openpyxl 未安装，无法解析 Excel 文件")
    
    result = []
    wb = load_workbook(file_path, read_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []
        
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
            if row_text.strip():
                rows.append(row_text)
        
        result.append({
            "sheet_name": sheet_name,
            "content": "\n".join(rows),
            "rows": rows,
            "filename": os.path.basename(file_path)
        })
    
    wb.close()
    return result


def _parse_text(file_path: str) -> List[Dict[str, Any]]:
    """
    解析纯文本文件
    
    Args:
        file_path: 文本文件路径
        
    Returns:
        包含文件内容的列表
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    paragraphs = [p.strip() for p in content.split('\n\n') if p.strip()]
    
    return [{
        "section_num": 1,
        "content": content,
        "paragraphs": paragraphs,
        "filename": os.path.basename(file_path)
    }]


def _parse_markdown(file_path: str) -> List[Dict[str, Any]]:
    """
    解析 Markdown 文件
    
    Args:
        file_path: Markdown 文件路径
        
    Returns:
        包含文件内容的列表
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 按标题分割
    sections = []
    current_section = {"title": "引言", "content": []}
    
    lines = content.split('\n')
    for line in lines:
        if line.startswith('# '):
            if current_section["content"]:
                sections.append(current_section)
            current_section = {"title": line[2:].strip(), "content": []}
        elif line.strip():
            current_section["content"].append(line)
    
    if current_section["content"]:
        sections.append(current_section)
    
    return [{
        "section_num": 1,
        "content": content,
        "sections": sections,
        "filename": os.path.basename(file_path)
    }]


def _extract_rules_from_content(
    parsed_data: List[Dict[str, Any]],
    focus_areas: Optional[List[str]] = None
) -> str:
    """
    生成规则抽取的提示词（供智能体调用内置 LLM 使用）
    
    Args:
        parsed_data: 解析后的文档数据
        focus_areas: 重点关注的规则类型
        
    Returns:
        系统提示词和用户提示词（由智能体调用内置 LLM）
    """
    # 构建提示词
    focus_desc = "全部规则类型"
    if focus_areas:
        focus_desc = "、".join(focus_areas)
    
    system_prompt = f"""你是一位专业的政策文件分析专家，专门从基金项目申报文件中提取关键规则信息。

请从提供的文档内容中，提取以下类型的规则：
1. 申请资格条件：申请人资质、项目类型、资金额度、配套要求、申报限制等
2. 审查标准：评审流程、评分标准、否决条件、材料要求、评审专家组成等
3. 时间节点：申报截止时间、评审时间、公示时间、立项时间、拨款时间等
4. 其他重要信息：联系方式、咨询渠道、注意事项、特殊说明等

要求：
- 每条规则必须精确、完整，不能模糊表述
- 必须标注每条规则的具体来源（文件名称、页码/章节、段落）
- 如果原文没有明确信息，不要编造
- 用 JSON 格式输出，格式如下：
{{
  "文件信息": {{
    "文件名": "文件名",
    "解析时间": "日期"
  }},
  "抽取结果": {{
    "申请资格": [
      {{
        "规则内容": "规则具体内容",
        "来源": {{
          "文件": "文件名",
          "页码": "页码或无页码",
          "章节": "章节名"
        }}
      }}
    ],
    "审查标准": [...],
    "时间节点": [...]
  }},
  "统计信息": {{
    "总规则数": 数字,
    "申请资格": 数字,
    "审查标准": 数字,
    "时间节点": 数字
  }}
}}
"""

    # 合并文档内容
    doc_contents = []
    for idx, data in enumerate(parsed_data):
        filename = data.get("filename", f"文件{idx+1}")
        content = data.get("content", "")
        doc_contents.append(f"[文件：{filename}]\n{content}")
    
    full_content = "\n\n---\n\n".join(doc_contents)
    
    user_prompt = f"""请从以下文档内容中提取{focus_desc}：

{full_content}

请严格按照上述 JSON 格式输出，不要添加任何额外说明。"""
    
    # 返回提示词，供智能体调用内置 LLM
    return f"系统提示词：\n{system_prompt}\n\n用户提示词：\n{user_prompt}"


@tool
async def parse_document(
    file_path: str,
    session_dir: Annotated[str, InjectedToolArg]
) -> str:
    """
    解析单个或多个文档文件，提取文本内容和结构信息
    
    支持格式：PDF、Word (.docx)、Excel (.xlsx)、TXT、Markdown
    
    Args:
        file_path: 文件路径（单个文件）或文件列表的 JSON 字符串
        
    Returns:
        解析后的结构化数据（JSON 格式）
    """
    try:
        # 判断是单个文件还是多个文件
        if file_path.startswith('['):
            files = json.loads(file_path)
        else:
            files = [file_path]
        
        all_results = []
        
        for f_path in files:
            # 处理相对路径
            if not os.path.isabs(f_path):
                f_path = os.path.join(session_dir, f_path)
            
            if not os.path.exists(f_path):
                all_results.append({
                    "error": f"文件不存在：{f_path}",
                    "filename": os.path.basename(f_path)
                })
                continue
            
            ext = os.path.splitext(f_path)[1].lower()
            
            if ext == '.pdf':
                result = _parse_pdf(f_path)
            elif ext in ['.docx', '.doc']:
                result = _parse_docx(f_path)
            elif ext in ['.xlsx', '.xls']:
                result = _parse_xlsx(f_path)
            elif ext == '.txt':
                result = _parse_text(f_path)
            elif ext == '.md':
                result = _parse_markdown(f_path)
            else:
                all_results.append({
                    "error": f"不支持的文件格式：{ext}",
                    "filename": os.path.basename(f_path)
                })
                continue
            
            all_results.extend(result)
        
        return json.dumps(all_results, ensure_ascii=False, indent=2)
        
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@tool
async def parse_and_prepare_rules(
    file_paths: str,
    focus_areas: Optional[str] = None,
    session_dir: Annotated[str, InjectedToolArg] = None
) -> str:
    """
    解析文档并生成规则抽取提示词
    
    此工具完成两步工作：
    1. 解析上传的文档文件（PDF/Word/Excel/TXT/Markdown）
    2. 生成规则抽取的系统提示词和用户提示词
    
    **智能体调用此工具后，会直接使用内置大模型处理提示词，用户无需配置任何 API。**
    
    Args:
        file_paths: 文件路径（单个文件）或文件列表的 JSON 字符串
        focus_areas: 重点关注的规则类型，逗号分隔（如"申请资格，审查标准"），为空则抽取全部
        
    Returns:
        包含解析后的文档内容和抽取提示词的字符串，智能体将据此调用内置 LLM 完成规则抽取
    """
    try:
        # 第一步：解析文档
        parse_result = await parse_document.run(file_paths)
        
        if "error" in parse_result:
            return f"文档解析失败：{parse_result['error']}"
        
        parsed_data = json.loads(parse_result)
        
        # 第二步：生成提示词
        focus_list = None
        if focus_areas:
            focus_list = [area.strip() for area in focus_areas.split(',') if area.strip()]
        
        prompts = _extract_rules_from_content(parsed_data, focus_list)
        
        # 构建输出，告诉智能体下一步该做什么
        output = f"""=== 文档解析完成 ===

已解析 {len(parsed_data)} 个文件：
"""
        for i, doc in enumerate(parsed_data, 1):
            output += f"{i}. {doc.get('filename', '未知文件')}\n"
        
        output += f"\n=== 规则抽取提示词 ===\n{prompts}\n\n=== 下一步操作 ===\n请调用内置大模型，使用上述系统提示词和用户提示词，提取基金项目申报规则，并按指定 JSON 格式输出结果。"
        
        return output
        
    except Exception as e:
        return f"准备失败：{str(e)}"


@tool
async def extract_rules_from_files(
    file_paths: str,
    focus_areas: Optional[str] = None,
    session_dir: Annotated[str, InjectedToolArg] = None
) -> str:
    """
    一站式工具：从文件直接准备规则抽取
    
    此工具完成以下工作：
    1. 解析上传的文档文件（PDF/Word/Excel/TXT/Markdown）
    2. 生成规则抽取的提示词
    
    **智能体调用此工具后，会直接使用内置大模型处理，用户无需配置任何 API。**
    
    Args:
        file_paths: 文件路径（单个文件）或文件列表的 JSON 字符串
        focus_areas: 重点关注的规则类型，逗号分隔
        
    Returns:
        包含解析结果和提示词的字符串，智能体将据此调用内置 LLM 完成抽取
    """
    return await parse_and_prepare_rules.run(file_paths, focus_areas, session_dir)


__all__ = [
    "parse_document",
    "parse_and_prepare_rules",
    "extract_rules_from_files",
]
